import pandas as pd
import requests
import io
import datetime
import os
import sys
# openpyxl 라이브러리 필수
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- [설정] 구글 스프레드시트 URL ---
SHEET_URLS = {
    'early': "https://docs.google.com/spreadsheets/d/1I_Cy5TZEnG0GmoThLPJJR7ZrXxUgXzsDDzu2zOtmjQI/export?format=csv&gid=214657398",
    'late': "https://docs.google.com/spreadsheets/d/1I_Cy5TZEnG0GmoThLPJJR7ZrXxUgXzsDDzu2zOtmjQI/export?format=csv&gid=1675631175"
}

class MokilReportGenerator:
    def __init__(self, mode):
        self.mode = mode
        self.raw_df = None
        self.classes = {i: {'g1': [], 'g2': [], 'g3': [], 'g4': []} for i in range(1, 16)}
        self.counts = {'g1': 0, 'g2': 0, 'g3': 0, 'g4': 0}
        self.report_date = "" 
        
        if mode == 'early':
            self.title = "2026학년도 목일중 전기고 진학 현황"
            self.groups = [
                {'id': 'g1', 'label': '영재학교', 'kwd': ['영재'], 'has_dept': False},
                {'id': 'g2', 'label': '과학고', 'kwd': ['과학고', '과고'], 'has_dept': False},
                {'id': 'g3', 'label': '예체미고', 'kwd': ['예고', '예술'], 'has_dept': False},
                {'id': 'g4', 'label': '특성화고', 'kwd': ['특성'], 'has_dept': True}
            ]
        else:
            self.title = "2026학년도 목일중 후기고 진학 현황"
            self.groups = [
                {'id': 'g1', 'label': '자사고', 'kwd': ['자사'], 'has_dept': False},
                {'id': 'g2', 'label': '외고/국제고', 'kwd': ['외고', '국제'], 'has_dept': False},
                {'id': 'g3', 'label': '비평준화고 / 중점학교', 'kwd': ['비평준', '중점'], 'has_dept': False}
            ]

    def set_date(self):
        print("-" * 50)
        if self.mode == 'early':
            default_date = "2025. 12. 3."
            prompt_msg = f"⚡ [전기고] 기준일 입력 (Enter = 기본값 '{default_date}'): "
        else:
            default_date = datetime.date.today().strftime("%Y. %m. %d.")
            prompt_msg = f"🍂 [후기고] 기준일 입력 (Enter = 오늘날짜 '{default_date}'): "
        
        try: user_input = input(prompt_msg).strip()
        except: user_input = ""
        
        self.report_date = user_input if user_input else default_date
        print(f"   👉 기준일 설정 완료: {self.report_date}")
        print("-" * 50)

    def fetch_google_sheet(self):
        url = SHEET_URLS[self.mode]
        print(f"📥 [{self.mode.upper()}] 데이터 다운로드 중...", end=" ", flush=True)
        try:
            response = requests.get(url)
            response.raise_for_status()
            self.raw_df = pd.read_csv(io.StringIO(response.content.decode('utf-8')), header=None)
            print("완료!")
            return True
        except Exception as e:
            print(f"\n❌ [오류] 데이터 다운로드 실패: {e}")
            return False

    def find_column_indices(self):
        df = self.raw_df
        header_row_idx = -1
        for i, row in df.iterrows():
            row_str = " ".join([str(x) for x in row.values])
            if "이름" in row_str or "성명" in row_str:
                header_row_idx = i
                break
        
        if header_row_idx == -1: return None
        header_row = df.iloc[header_row_idx]
        
        name_cols = []
        for idx, val in header_row.items():
            if "이름" in str(val) or "성명" in str(val): name_cols.append(idx)
        
        group_indices = {}
        for i, group in enumerate(self.groups):
            target_idx = -1
            if self.mode == 'early':
                offset = 1 if len(name_cols) == 3 else 0
                if i >= offset and (i - offset) < len(name_cols): target_idx = name_cols[i - offset]
            else:
                if i < len(name_cols): target_idx = name_cols[i]
            
            if target_idx != -1: group_indices[group['id']] = self._detect_columns(target_idx, header_row)
            else: group_indices[group['id']] = {'name': -1}
                
        return header_row_idx, group_indices

    def _detect_columns(self, name_idx, header_row):
        info = {'name': name_idx, 'class': name_idx - 1, 'gender': name_idx + 1, 'school': name_idx + 2, 'dept': name_idx + 3, 'pass': -1}
        for offset in range(1, 7):
            check_idx = info['school'] + offset
            if check_idx in header_row.index:
                val = str(header_row[check_idx])
                if any(x in val for x in ['합', '불', '당락', '합격', '결과']):
                    info['pass'] = check_idx
                    break
        return info

    def process(self):
        self.set_date()
        if not self.fetch_google_sheet(): return
        result = self.find_column_indices()
        if not result: return

        h_idx, indices = result
        df = self.raw_df.iloc[h_idx+1:] 

        for _, row in df.iterrows():
            for group in self.groups:
                gid = group['id']
                idx = indices[gid]
                if idx['name'] == -1 or pd.isna(row[idx['name']]): continue
                cls_num = self._parse_class(str(row[idx['class']]))
                if not cls_num: continue
                if idx['pass'] != -1:
                    pass_val = str(row[idx['pass']]).strip()
                    if "합" not in pass_val: continue

                school_name = str(row[idx['school']]).strip()
                if self.mode == 'early' and gid == 'g3': school_name = self._clean_arts_school(school_name)
                else: school_name = school_name.split('(')[0]

                student = {'name': str(row[idx['name']]).strip(), 'gender': '남' if '남' in str(row[idx['gender']]) else '여', 'school': school_name, 'dept': str(row[idx['dept']]).strip() if group['has_dept'] else ''}
                self.classes[cls_num][gid].append(student)
                self.counts[gid] += 1

        self.save_html()
        self.save_excel()

    def _parse_class(self, val):
        import re
        if '-' in val: return int(val.split('-')[1])
        nums = re.findall(r'\d+', val)
        return int(nums[-1]) if nums else None

    def _clean_arts_school(self, name):
        name = name.split('(')[0].strip()
        majors = ['미술', '음악', '무용', '연극', '영화', '성악', '작곡', '디자인', '만화']
        for m in majors:
            idx = name.find(m)
            if idx > 1: return name[:idx].strip()
        return name.split(' ')[0]

    def save_html(self):
        visible_groups = [g for g in self.groups if self.counts[g['id']] > 0]
        
        # 헤더 생성 (검색창 포함)
        thead1 = '<tr><th rowspan="3" class="thick-right" style="width:50px;">학반</th>'
        thead2 = '<tr>'
        thead3 = '<tr>' # 검색 필터 행 추가
        
        for g in visible_groups:
            cols = 4 if g['has_dept'] else 3
            thead1 += f'<th colspan="{cols}" class="bg-group thick-right">{g["label"]}</th>'
            thead2 += '<th style="width:60px;">이름</th><th style="width:40px;">성별</th>'
            
            # 검색창 셀 생성 (colspan 적용)
            thead3 += f'<th colspan="{cols}" class="filter-cell thick-right"><input type="text" class="col-filter" data-group="{g["id"]}" placeholder="{g["label"]} 검색" onkeyup="applyColumnFilter()"></th>'

            if g['has_dept']: thead2 += '<th>학교명</th><th class="thick-right">학과</th>'
            else: thead2 += '<th class="thick-right">학교명</th>'
            
        thead1 += '</tr>'; thead2 += '</tr>'; thead3 += '</tr>'

        tbody = ''
        stats = {g['id']: {'m':0, 'f':0, 'schools':{}} for g in self.groups}
        
        for i in range(1, 16):
            c_data = self.classes[i]
            row_counts = [len(c_data[g['id']]) for g in visible_groups]
            max_rows = max(row_counts) if row_counts else 0
            if max_rows == 0: max_rows = 1
            
            for r in range(max_rows):
                cls_border = 'thick-top' if r == 0 else ''
                row_cells_html = ""
                
                for g in visible_groups:
                    st_list = c_data[g['id']]
                    if r < len(st_list):
                        s = st_list[r]
                        stats[g['id']]['m' if s['gender']=='남' else 'f'] += 1
                        sch = s['school']
                        stats[g['id']]['schools'][sch] = stats[g['id']]['schools'].get(sch, 0) + 1
                        
                        # 데이터 속성 추가 (그룹별 검색용)
                        data_attrs = f'data-group-{g["id"]}-school="{s["school"]}" data-group-{g["id"]}-name="{s["name"]}"'
                        
                        row_cells_html += f'<td class="{cls_border} col-name" {data_attrs}>{s["name"]}</td><td class="{cls_border} col-gender" {data_attrs}>{s["gender"]}</td>'
                        if g['has_dept']: row_cells_html += f'<td class="{cls_border} col-school" {data_attrs}>{s["school"]}</td><td class="{cls_border} thick-right" {data_attrs}>{s["dept"]}</td>'
                        else: row_cells_html += f'<td class="{cls_border} thick-right col-school" {data_attrs}>{s["school"]}</td>'
                    else:
                        # 빈 셀 (검색 대상 아님)
                        row_cells_html += f'<td class="{cls_border}"></td><td class="{cls_border}"></td>'
                        if g['has_dept']: row_cells_html += f'<td class="{cls_border}"></td><td class="{cls_border} thick-right"></td>'
                        else: row_cells_html += f'<td class="{cls_border} thick-right"></td>'

                tbody += '<tr>'
                if r == 0: tbody += f'<td rowspan="{max_rows}" class="{cls_border} thick-right font-bold class-cell">3-{i}</td>'
                tbody += row_cells_html + '</tr>'

        tfoot = f'<tfoot><tr class="thick-top bg-gray-50 font-bold"><td class="thick-right">남</td>'
        for g in visible_groups:
            col = 4 if g['has_dept'] else 3
            tfoot += f'<td colspan="{col}" class="thick-right">{stats[g["id"]]["m"]}명</td>'
        tfoot += '</tr><tr class="bg-gray-50 font-bold"><td class="thick-right">여</td>'
        for g in visible_groups:
            col = 4 if g['has_dept'] else 3
            tfoot += f'<td colspan="{col}" class="thick-right">{stats[g["id"]]["f"]}명</td>'
        tfoot += '</tr><tr class="thick-top bg-group font-bold border-b-2 border-black"><td class="thick-right">계</td>'
        for g in visible_groups:
            col = 4 if g['has_dept'] else 3
            tfoot += f'<td colspan="{col}" class="thick-right">{stats[g["id"]]["m"] + stats[g["id"]]["f"]}명</td>'
        tfoot += '</tr></tfoot>'

        summary_html = '<div class="stats-container"><div class="stats-header">통계 요약</div>'
        total_all, total_m, total_f = 0, 0, 0
        for g in self.groups:
            st = stats[g['id']]
            sub_tot = st['m'] + st['f']
            total_all += sub_tot; total_m += st['m']; total_f += st['f']
            sch_str = ", ".join([f"{k}: {v}명" for k, v in sorted(st['schools'].items())])
            summary_html += f'<div class="stats-row"><span class="stats-label">{g["label"]}:</span> <span>총 {sub_tot}명 (남: {st["m"]}명, 여: {st["f"]}명)</span>'
            if sch_str: summary_html += f'<div class="stats-school-list">└ {sch_str}</div>'
            summary_html += '</div>'
        summary_html += f'<div class="stats-total-box">전체 합격 인원: 총 {total_all}명 (남: {total_m}명, 여: {total_f}명)</div></div>'

        output_dir = "reports"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        filename = os.path.join(output_dir, f"목일중_{self.mode}_진학현황.html")
        full_html = f"""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><title>{self.title}</title><style>
        body {{ font-family: 'Malgun Gothic', 'Noto Sans KR', sans-serif; padding: 30px; background: #f9fafb; }}
        .container {{ max-width: 1600px; margin: 0 auto; background: white; padding: 40px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-radius: 8px; }}
        .print-hide {{ }} @media print {{ .print-hide, .filter-cell {{ display: none !important; }} body {{ padding: 0; background: white; }} .container {{ box-shadow: none; padding: 0; }} }}
        table {{ width: 100%; border-collapse: collapse; text-align: center; border: 2px solid #000; font-size: 10pt; }}
        th, td {{ border: 1px solid #000; padding: 5px 2px; vertical-align: middle; white-space: nowrap; }}
        thead th {{ background-color: #f8f9fa; font-weight: bold; border-bottom: 1px solid #000; height: 30px; }}
        .filter-cell {{ background-color: #e9ecef; padding: 4px; }}
        .col-filter {{ width: 90%; padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 0.9em; text-align: center; }}
        .bg-group {{ background-color: #e9ecef !important; border-bottom: 2px solid #000 !important; }}
        .thick-top {{ border-top: 2px solid #000 !important; }}
        .thick-right {{ border-right: 2px solid #000 !important; }}
        .stats-container {{ margin-top: 30px; border-top: 2px solid #000; padding-top: 20px; font-size: 11pt; line-height: 1.6; }}
        .stats-header {{ font-size: 13pt; font-weight: bold; text-decoration: underline; margin-bottom: 15px; }}
        .stats-row {{ margin-bottom: 8px; }}
        .stats-label {{ display: inline-block; font-weight: bold; width: 160px; }}
        .stats-school-list {{ margin-left: 10px; color: #444; font-size: 10pt; }}
        .stats-total-box {{ margin-top: 20px; padding-top: 15px; border-top: 1px solid #aaa; font-weight: bold; font-size: 12pt; }}
        .hidden-cell {{ color: transparent; user-select: none; }} /* 텍스트만 숨김 */
        </style></head><body><div class="container">
        <h2 style="text-align:center; font-weight:bold; margin-bottom: 20px;">{self.title}</h2>
        <p style="text-align:right; font-size:10pt; margin-bottom: 5px;">(기준: {self.report_date} 최종 합불)</p>
        
        <table id="dataTable"><thead>{thead1}{thead2}{thead3}</thead><tbody>{tbody}</tbody>{tfoot}</table>
        {summary_html}</div>
        
        <script>
        function applyColumnFilter() {{
            const filters = document.querySelectorAll('.col-filter');
            const rows = document.querySelectorAll('tbody tr');
            
            // 각 행마다 검사
            rows.forEach(row => {{
                // 각 필터(학교군)마다 검사
                filters.forEach(filter => {{
                    const groupId = filter.getAttribute('data-group');
                    const keyword = filter.value.toLowerCase();
                    
                    // 해당 행, 해당 그룹의 데이터 셀들 찾기 (school, name, gender 등)
                    // 데이터 속성 활용: data-group-g1-school="..."
                    const cells = row.querySelectorAll(`[data-group-${{groupId}}-school]`);
                    
                    cells.forEach(cell => {{
                        if (!keyword) {{
                            cell.style.opacity = '1'; // 필터 없으면 보임
                            return;
                        }}
                        
                        const schoolVal = cell.getAttribute(`data-group-${{groupId}}-school`).toLowerCase();
                        const nameVal = cell.getAttribute(`data-group-${{groupId}}-name`).toLowerCase();
                        
                        // 학교명이나 이름에 키워드가 포함되면 보임, 아니면 숨김(투명화)
                        if (schoolVal.includes(keyword) || nameVal.includes(keyword)) {{
                            cell.style.opacity = '1';
                        }} else {{
                            cell.style.opacity = '0.1'; // 흐리게 처리 (완전 숨기면 표 깨짐 방지)
                        }}
                    }});
                }});
            }});
        }}
        </script></body></html>"""
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(full_html)
        print(f"✅ [{self.mode.upper()}] HTML 파일 생성 완료: {os.path.abspath(filename)}")

    def save_excel(self):
        output_dir = "reports"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        filename = os.path.join(output_dir, f"목일중_{self.mode}_진학현황.xlsx")
        visible_groups = [g for g in self.groups if self.counts[g['id']] > 0]
        data_rows = []
        
        header1 = ["학반"]
        for g in visible_groups:
            cols = 4 if g['has_dept'] else 3
            header1.append(g['label']); header1.extend([""] * (cols - 1))
        data_rows.append(header1)

        header2 = [""]
        for g in visible_groups:
            header2.extend(["이름", "성별", "학교명"])
            if g['has_dept']: header2.append("학과")
        data_rows.append(header2)

        merge_info = []
        current_row = 3

        for i in range(1, 16):
            c_data = self.classes[i]
            row_counts = [len(c_data[g['id']]) for g in visible_groups]
            max_rows = max(row_counts) if row_counts else 0
            if max_rows == 0: max_rows = 1
            
            merge_info.append((current_row, current_row + max_rows - 1, 1, f"3-{i}"))

            for r in range(max_rows):
                row_data = [""] 
                for g in visible_groups:
                    st_list = c_data[g['id']]
                    if r < len(st_list):
                        s = st_list[r]
                        row_data.extend([s['name'], s['gender'], s['school']])
                        if g['has_dept']: row_data.append(s['dept'])
                    else:
                        row_data.extend(["", "", ""])
                        if g['has_dept']: row_data.append("")
                data_rows.append(row_data)
            current_row += max_rows

        df_excel = pd.DataFrame(data_rows)

        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_excel.to_excel(writer, sheet_name='Sheet1', header=False, index=False)
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center', vertical='center')
                header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                header_font = Font(name='맑은 고딕', size=10, bold=True)
                base_font = Font(name='맑은 고딕', size=10)

                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border; cell.alignment = center_align; cell.font = base_font
                        if cell.row <= 2: cell.fill = header_fill; cell.font = header_font

                col_idx = 2
                for g in visible_groups:
                    cols = 4 if g['has_dept'] else 3
                    if cols > 1:
                        worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + cols - 1)
                        worksheet.cell(row=1, column=col_idx).value = g['label']
                    col_idx += cols

                for r_start, r_end, c_idx, val in merge_info:
                    if r_start == r_end: worksheet.cell(row=r_start, column=c_idx).value = val
                    else:
                        worksheet.merge_cells(start_row=r_start, start_column=c_idx, end_row=r_end, end_column=c_idx)
                        worksheet.cell(row=r_start, column=c_idx).value = val

                worksheet.column_dimensions['A'].width = 8
                curr_col = 2
                for g in visible_groups:
                    worksheet.column_dimensions[get_column_letter(curr_col)].width = 10
                    worksheet.column_dimensions[get_column_letter(curr_col+1)].width = 6
                    worksheet.column_dimensions[get_column_letter(curr_col+2)].width = 18
                    if g['has_dept']:
                        worksheet.column_dimensions[get_column_letter(curr_col+3)].width = 18
                        curr_col += 4
                    else: curr_col += 3

            print(f"✅ [{self.mode.upper()}] 엑셀 파일 생성 완료: {os.path.abspath(filename)}")
        except Exception as e:
            print(f"❌ 엑셀 저장 실패: {e}")

if __name__ == "__main__":
    print("=== 목일중 진학 현황 자동 생성기 (V22: Independent Filter) ===")
    MokilReportGenerator('early').process()
    print("\n" + "-"*50 + "\n")
    MokilReportGenerator('late').process()